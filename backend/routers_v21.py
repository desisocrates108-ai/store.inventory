"""V2.1 router — additive endpoints for OCR alias learning, franchise tiers,
multi-source ordering, bulk inventory import, editable POs, date filtering.

All routes are mounted under /api by server.py.
Existing endpoints are NOT modified — these are *additions* preserving full
backward compatibility.
"""
from __future__ import annotations

import io
import json
import logging
import os
from datetime import datetime, timezone
from typing import Optional, List

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from auth_utils import require_roles, get_current_user
from models import (
    FranchiseTier, CategoryMarginOverride, StarterKitItem, OcrAlias, Indent, IndentLineItem,
    PurchaseOrder, POLineItem, gen_id, now_iso,
)

logger = logging.getLogger(__name__)

router = APIRouter()


# ============================================================
#  Helpers (db handle is injected via dependency from server.py)
# ============================================================
_db = None
_log_audit = None
_gen_sequence = None
_save_uploaded_file = None
_adjust_stock = None
_ocr_service = None
_upload_dir = None


def init(db, log_audit_fn, gen_sequence_fn, save_uploaded_file_fn, adjust_stock_fn, ocr_service_mod, upload_dir):
    """Called once from server.py at import-time to wire dependencies."""
    global _db, _log_audit, _gen_sequence, _save_uploaded_file, _adjust_stock, _ocr_service, _upload_dir
    _db = db
    _log_audit = log_audit_fn
    _gen_sequence = gen_sequence_fn
    _save_uploaded_file = save_uploaded_file_fn
    _adjust_stock = adjust_stock_fn
    _ocr_service = ocr_service_mod
    _upload_dir = upload_dir


# ============================================================
#  1. FRANCHISE TIER PRICING
# ============================================================
class FranchiseTierIn(BaseModel):
    name: str
    margin_percent: float = 22.0
    category_overrides: List[CategoryMarginOverride] = []
    color: str = ""
    active: bool = True


@router.get("/franchise-tiers", tags=["pricing"])
async def list_tiers(user: dict = Depends(get_current_user)):
    docs = await _db.franchise_tiers.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    return docs


@router.post("/franchise-tiers", tags=["pricing"])
async def create_tier(body: FranchiseTierIn, request: Request,
                      user: dict = Depends(require_roles("super_admin", "hub_accountant"))):
    if not body.name.strip():
        raise HTTPException(400, "Tier name required")
    existing = await _db.franchise_tiers.find_one({"name": body.name.strip()}, {"_id": 0})
    if existing:
        raise HTTPException(409, f"Tier '{body.name}' already exists")
    tier = FranchiseTier(
        name=body.name.strip(),
        margin_percent=body.margin_percent,
        category_overrides=body.category_overrides,
        color=body.color,
        active=body.active,
        is_system=False,
    )
    await _db.franchise_tiers.insert_one(tier.model_dump())
    await _log_audit(user, "tier.create", "franchise_tier", tier.id,
                     after={"name": tier.name, "margin": tier.margin_percent}, request=request)
    return tier.model_dump()


@router.put("/franchise-tiers/{tier_id}", tags=["pricing"])
async def update_tier(tier_id: str, body: FranchiseTierIn, request: Request,
                      user: dict = Depends(require_roles("super_admin", "hub_accountant"))):
    tier = await _db.franchise_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        raise HTTPException(404, "Tier not found")
    update = {
        "name": body.name.strip(),
        "margin_percent": body.margin_percent,
        "category_overrides": [c.model_dump() for c in body.category_overrides],
        "color": body.color,
        "active": body.active,
    }
    await _db.franchise_tiers.update_one({"id": tier_id}, {"$set": update})
    await _log_audit(user, "tier.update", "franchise_tier", tier_id,
                     before={"margin": tier.get("margin_percent")},
                     after={"margin": body.margin_percent}, request=request)
    return {"ok": True}


@router.delete("/franchise-tiers/{tier_id}", tags=["pricing"])
async def delete_tier(tier_id: str, request: Request,
                      user: dict = Depends(require_roles("super_admin"))):
    tier = await _db.franchise_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        raise HTTPException(404, "Tier not found")
    if tier.get("is_system"):
        raise HTTPException(403, "Cannot delete system tier")
    # Unassign tier from any franchises that used it
    await _db.franchises.update_many({"tier_id": tier_id}, {"$set": {"tier_id": None}})
    await _db.franchise_tiers.delete_one({"id": tier_id})
    await _log_audit(user, "tier.delete", "franchise_tier", tier_id,
                     before={"name": tier["name"]}, request=request)
    return {"ok": True}


class AssignTierRequest(BaseModel):
    tier_id: Optional[str] = None
    replace_starter_kit: bool = False  # V2.4 — if true, capture latest snapshot from new tier


@router.put("/franchises/{fid}/tier", tags=["pricing"])
async def assign_tier(fid: str, body: AssignTierRequest, request: Request,
                      user: dict = Depends(require_roles("super_admin", "hub_accountant"))):
    fr = await _db.franchises.find_one({"id": fid}, {"_id": 0})
    if not fr:
        raise HTTPException(404, "Franchise not found")
    tier_doc: Optional[dict] = None
    if body.tier_id:
        tier_doc = await _db.franchise_tiers.find_one({"id": body.tier_id}, {"_id": 0})
        if not tier_doc:
            raise HTTPException(404, "Tier not found")
    update: dict = {"tier_id": body.tier_id}
    snapshot_action = "kept"
    if body.tier_id and body.replace_starter_kit and tier_doc:
        snap = _capture_snapshot(tier_doc)
        # Preserve existing snapshot's version counter
        prev_version = int((fr.get("starter_kit") or {}).get("version", 0) or 0)
        snap["version"] = prev_version + 1
        update["starter_kit"] = snap
        snapshot_action = "replaced"
    elif body.tier_id and not fr.get("starter_kit") and tier_doc:
        # First-time tier assignment: always capture, since there's nothing to overwrite.
        update["starter_kit"] = _capture_snapshot(tier_doc)
        snapshot_action = "captured"
    await _db.franchises.update_one({"id": fid}, {"$set": update})
    await _log_audit(user, "franchise.tier_assign", "franchise", fid,
                     before={"tier_id": fr.get("tier_id")},
                     after={"tier_id": body.tier_id, "snapshot": snapshot_action},
                     request=request)
    return {"ok": True, "snapshot": snapshot_action}


def _resolve_margin(tier: Optional[dict], category: str, default: float = 22.0) -> float:
    """Compute effective margin for (tier, category)."""
    if not tier:
        return default
    for ov in tier.get("category_overrides", []) or []:
        if (ov.get("category") or "").strip().lower() == (category or "").strip().lower():
            return float(ov.get("margin_percent", tier.get("margin_percent", default)))
    return float(tier.get("margin_percent", default))


@router.get("/franchise-tiers/{tier_id}/preview", tags=["pricing"])
async def preview_tier_pricing(tier_id: str, user: dict = Depends(get_current_user)):
    """Return tier-aware product prices for preview/UI use."""
    tier = await _db.franchise_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        raise HTTPException(404, "Tier not found")
    products = await _db.products.find({"active": True}, {"_id": 0}).to_list(5000)
    rows = []
    for p in products:
        landing = p.get("landing_price", 0) or 0
        margin = _resolve_margin(tier, p.get("category", ""), p.get("margin_percent", 22))
        price = round(landing * (1 + margin / 100), 2)
        rows.append({
            "product_id": p["id"],
            "sku": p["sku"],
            "name": p["name"],
            "category": p.get("category", ""),
            "landing_price": landing,
            "margin_percent": margin,
            "tier_price": price,
            "mrp": p.get("mrp", 0),
        })
    return {"tier": tier, "rows": rows}


# ---- Starter-Kit Template (per-tier recommended product list with per-item pricing) ----
class StarterKitItemIn(BaseModel):
    product_id: str
    recommended_qty: float = 1.0
    discount_percent: float = 0.0
    margin_percent: Optional[float] = None  # None → fall back to tier.margin_percent


class StarterKitIn(BaseModel):
    default_margin_percent: Optional[float] = None  # tier-level fallback (used when item's margin_percent is null)
    items: List[StarterKitItemIn] = []


def _resolve_kit_item_pricing(item: dict, product: dict, tier_margin: float) -> dict:
    """Return the per-row pricing fields (cost / margin / discount / selling / line totals)
    given a raw saved item dict + product doc + tier-level margin fallback.

    Pure function — used by the tier preview, the franchise snapshot capture,
    and the New-Franchise-Setup indent prefill. Keeping it in one place is the
    'reusable Pricing Engine' the user asked for."""
    cost = float(product.get("landing_price", 0) or 0)
    mrp = float(product.get("mrp", 0) or 0)
    item_margin = item.get("margin_percent")
    margin = float(item_margin) if item_margin is not None else float(tier_margin or 0)
    base_price = round(cost * (1 + margin / 100), 2)
    discount = float(item.get("discount_percent", 0) or 0)
    selling = round(base_price * (1 - discount / 100), 2)
    qty = float(item.get("recommended_qty", 0) or 0)
    return {
        "cost_price": cost,
        "mrp": mrp,
        "margin_percent_effective": margin,
        "base_price": base_price,            # franchise price BEFORE discount
        "discount_percent": discount,
        "selling_price": selling,            # AFTER discount
        "recommended_qty": qty,
        "inventory_value": round(cost * qty, 2),
        "selling_value": round(selling * qty, 2),
        "expected_profit": round((selling - cost) * qty, 2),
    }


@router.get("/franchise-tiers/{tier_id}/starter-kit", tags=["pricing"])
async def get_starter_kit(tier_id: str, user: dict = Depends(get_current_user)):
    """Return the saved Starter-Kit template for a tier.

    Items are returned enriched with sku/name + per-item live pricing so the UI
    can render a single row without a second roundtrip to /products."""
    tier = await _db.franchise_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        raise HTTPException(404, "Tier not found")
    tier_margin = float(tier.get("margin_percent", 22.0) or 22.0)
    items = tier.get("starter_kit_items", []) or []
    enriched: list = []
    if items:
        pids = [i.get("product_id") for i in items if i.get("product_id")]
        products = await _db.products.find({"id": {"$in": pids}}, {"_id": 0}).to_list(2000)
        pmap = {p["id"]: p for p in products}
        for it in items:
            pid = it.get("product_id")
            p = pmap.get(pid) or {}
            pricing = _resolve_kit_item_pricing(it, p, tier_margin)
            enriched.append({
                "product_id": pid,
                "sku": it.get("sku") or p.get("sku", ""),
                "name": it.get("name") or p.get("name", ""),
                "category": p.get("category", ""),
                "recommended_qty": pricing["recommended_qty"],
                "discount_percent": pricing["discount_percent"],
                # margin_percent is stored as null if user wants tier fallback;
                # we expose both the override and the effective resolved value.
                "margin_percent": it.get("margin_percent"),
                "margin_percent_effective": pricing["margin_percent_effective"],
                "cost_price": pricing["cost_price"],
                "mrp": pricing["mrp"],
                "base_price": pricing["base_price"],
                "selling_price": pricing["selling_price"],
                "inventory_value": pricing["inventory_value"],
                "selling_value": pricing["selling_value"],
                "expected_profit": pricing["expected_profit"],
            })
    return {
        "tier_id": tier_id,
        "tier_name": tier.get("name"),
        "default_margin_percent": tier_margin,
        "items": enriched,
    }


@router.put("/franchise-tiers/{tier_id}/starter-kit", tags=["pricing"])
async def save_starter_kit(tier_id: str, body: StarterKitIn, request: Request,
                            user: dict = Depends(require_roles("super_admin", "hub_accountant"))):
    tier = await _db.franchise_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        raise HTTPException(404, "Tier not found")

    # Deduplicate by product_id, drop empties / unknowns.
    seen: set = set()
    items: list = []
    for it in body.items or []:
        pid = (it.product_id or "").strip()
        if not pid or pid in seen:
            continue
        seen.add(pid)
        prod = await _db.products.find_one({"id": pid}, {"_id": 0})
        if not prod:
            continue  # silently drop stale ids
        items.append({
            "product_id": pid,
            "sku": prod.get("sku", ""),
            "name": prod.get("name", ""),
            "recommended_qty": float(it.recommended_qty or 0),
            "discount_percent": float(it.discount_percent or 0),
            "margin_percent": (float(it.margin_percent) if it.margin_percent is not None else None),
        })

    update: dict = {"starter_kit_items": items}
    if body.default_margin_percent is not None:
        update["margin_percent"] = float(body.default_margin_percent)

    await _db.franchise_tiers.update_one({"id": tier_id}, {"$set": update})
    await _log_audit(user, "tier.starter_kit.save", "franchise_tier", tier_id,
                     after={"item_count": len(items)}, request=request)
    return {"ok": True, "item_count": len(items)}


@router.get("/franchise-tiers/{tier_id}/starter-kit/preview", tags=["pricing"])
async def preview_starter_kit_totals(tier_id: str, user: dict = Depends(get_current_user)):
    """Aggregate rollup of a tier's Starter-Kit — total items, total qty,
    inventory value, expected selling value, expected profit. Used by the
    Franchise-Tiers UI 'Preview' cards."""
    tier = await _db.franchise_tiers.find_one({"id": tier_id}, {"_id": 0})
    if not tier:
        raise HTTPException(404, "Tier not found")
    tier_margin = float(tier.get("margin_percent", 22.0) or 22.0)
    items = tier.get("starter_kit_items", []) or []
    if not items:
        return {
            "tier_id": tier_id, "tier_name": tier.get("name"),
            "total_items": 0, "total_qty": 0.0,
            "inventory_value": 0.0, "selling_value": 0.0, "expected_profit": 0.0,
            "rows": [],
        }
    pids = [i.get("product_id") for i in items if i.get("product_id")]
    products = await _db.products.find({"id": {"$in": pids}}, {"_id": 0}).to_list(2000)
    pmap = {p["id"]: p for p in products}
    rows: list = []
    total_qty = inv_val = sell_val = profit = 0.0
    for it in items:
        p = pmap.get(it.get("product_id"), {})
        pricing = _resolve_kit_item_pricing(it, p, tier_margin)
        total_qty += pricing["recommended_qty"]
        inv_val += pricing["inventory_value"]
        sell_val += pricing["selling_value"]
        profit += pricing["expected_profit"]
        rows.append({
            "sku": it.get("sku") or p.get("sku", ""),
            "name": it.get("name") or p.get("name", ""),
            **pricing,
        })
    return {
        "tier_id": tier_id,
        "tier_name": tier.get("name"),
        "total_items": len(items),
        "total_qty": round(total_qty, 2),
        "inventory_value": round(inv_val, 2),
        "selling_value": round(sell_val, 2),
        "expected_profit": round(profit, 2),
        "rows": rows,
    }


def _capture_snapshot(tier: dict) -> dict:
    """Return a Starter-Kit snapshot dict ready to embed on a franchise document."""
    return {
        "tier_id": tier.get("id"),
        "tier_name": tier.get("name", ""),
        "default_margin_percent": float(tier.get("margin_percent", 22.0) or 22.0),
        "items": [
            {
                "product_id": it.get("product_id"),
                "sku": it.get("sku", ""),
                "name": it.get("name", ""),
                "recommended_qty": float(it.get("recommended_qty") or 0),
                "discount_percent": float(it.get("discount_percent") or 0),
                "margin_percent": (float(it.get("margin_percent")) if it.get("margin_percent") is not None else None),
            }
            for it in (tier.get("starter_kit_items") or [])
        ],
        "captured_at": now_iso(),
        "version": 1,
    }


@router.get("/franchises/{fid}/starter-kit", tags=["pricing"])
async def get_franchise_starter_kit(fid: str, user: dict = Depends(get_current_user)):
    """Per-franchise editable copy of a Starter-Kit snapshot.

    franchise_managers can read their own; others (admin/accountant) any.
    If the franchise has no snapshot yet, returns {has_snapshot: False}."""
    fr = await _db.franchises.find_one({"id": fid}, {"_id": 0})
    if not fr:
        raise HTTPException(404, "Franchise not found")
    if user["role"] == "franchise_manager" and user.get("franchise_id") != fid:
        raise HTTPException(403, "Forbidden")
    snap = fr.get("starter_kit") or None
    if not snap:
        return {"has_snapshot": False, "franchise_id": fid, "franchise_name": fr.get("name", "")}

    items = snap.get("items") or []
    tier_margin = float(snap.get("default_margin_percent", 22.0) or 22.0)
    pids = [i.get("product_id") for i in items if i.get("product_id")]
    products = await _db.products.find({"id": {"$in": pids}}, {"_id": 0}).to_list(2000) if pids else []
    pmap = {p["id"]: p for p in products}
    enriched: list = []
    for it in items:
        p = pmap.get(it.get("product_id"), {})
        pricing = _resolve_kit_item_pricing(it, p, tier_margin)
        enriched.append({
            "product_id": it.get("product_id"),
            "sku": it.get("sku") or p.get("sku", ""),
            "name": it.get("name") or p.get("name", ""),
            "category": p.get("category", ""),
            "recommended_qty": pricing["recommended_qty"],
            "discount_percent": pricing["discount_percent"],
            "margin_percent": it.get("margin_percent"),
            "margin_percent_effective": pricing["margin_percent_effective"],
            "cost_price": pricing["cost_price"],
            "mrp": pricing["mrp"],
            "base_price": pricing["base_price"],
            "selling_price": pricing["selling_price"],
            "inventory_value": pricing["inventory_value"],
            "selling_value": pricing["selling_value"],
            "expected_profit": pricing["expected_profit"],
        })
    return {
        "has_snapshot": True,
        "franchise_id": fid,
        "franchise_name": fr.get("name", ""),
        "tier_id": snap.get("tier_id"),
        "tier_name": snap.get("tier_name", ""),
        "default_margin_percent": tier_margin,
        "captured_at": snap.get("captured_at"),
        "version": snap.get("version", 1),
        "items": enriched,
    }


class FranchiseStarterKitIn(BaseModel):
    default_margin_percent: Optional[float] = None
    items: List[StarterKitItemIn] = []


@router.put("/franchises/{fid}/starter-kit", tags=["pricing"])
async def save_franchise_starter_kit(fid: str, body: FranchiseStarterKitIn, request: Request,
                                       user: dict = Depends(require_roles("super_admin", "hub_accountant"))):
    """Replace the per-franchise Starter-Kit snapshot with the posted items.
    Bumps the snapshot version. Does NOT touch the source tier template."""
    fr = await _db.franchises.find_one({"id": fid}, {"_id": 0})
    if not fr:
        raise HTTPException(404, "Franchise not found")

    seen: set = set()
    items: list = []
    for it in body.items or []:
        pid = (it.product_id or "").strip()
        if not pid or pid in seen:
            continue
        seen.add(pid)
        prod = await _db.products.find_one({"id": pid}, {"_id": 0})
        if not prod:
            continue
        items.append({
            "product_id": pid,
            "sku": prod.get("sku", ""),
            "name": prod.get("name", ""),
            "recommended_qty": float(it.recommended_qty or 0),
            "discount_percent": float(it.discount_percent or 0),
            "margin_percent": (float(it.margin_percent) if it.margin_percent is not None else None),
        })

    prev = fr.get("starter_kit") or {}
    snap = {
        "tier_id": prev.get("tier_id") or fr.get("tier_id"),
        "tier_name": prev.get("tier_name", ""),
        "default_margin_percent": (
            float(body.default_margin_percent) if body.default_margin_percent is not None
            else float(prev.get("default_margin_percent", 22.0) or 22.0)
        ),
        "items": items,
        "captured_at": now_iso(),
        "version": int(prev.get("version", 0) or 0) + 1,
    }
    await _db.franchises.update_one({"id": fid}, {"$set": {"starter_kit": snap}})
    await _log_audit(user, "franchise.starter_kit.save", "franchise", fid,
                     after={"item_count": len(items), "version": snap["version"]}, request=request)
    return {"ok": True, "item_count": len(items), "version": snap["version"]}


@router.get("/franchises/{fid}/starter-kit/indent-prefill", tags=["pricing"])
async def starter_kit_indent_prefill(fid: str, user: dict = Depends(get_current_user)):
    """Return Starter-Kit lines pre-shaped as Indent line items (server-computed
    selling_price + line_total). The UI loads these into the New-Order page
    when priority='new_franchise_setup' is selected. The lines remain editable
    in the UI; submitting goes through the normal POST /api/indents path."""
    fr = await _db.franchises.find_one({"id": fid}, {"_id": 0})
    if not fr:
        raise HTTPException(404, "Franchise not found")
    if user["role"] == "franchise_manager" and user.get("franchise_id") != fid:
        raise HTTPException(403, "Forbidden")
    snap = fr.get("starter_kit") or None
    if not snap:
        raise HTTPException(404, "No Starter-Kit snapshot on this franchise — assign a tier first.")
    tier_margin = float(snap.get("default_margin_percent", 22.0) or 22.0)
    items = snap.get("items") or []
    pids = [i.get("product_id") for i in items if i.get("product_id")]
    products = await _db.products.find({"id": {"$in": pids}}, {"_id": 0}).to_list(2000) if pids else []
    pmap = {p["id"]: p for p in products}
    lines: list = []
    for it in items:
        p = pmap.get(it.get("product_id"))
        if not p:
            continue  # stale snapshot — skip the line
        pricing = _resolve_kit_item_pricing(it, p, tier_margin)
        lines.append({
            "product_id": it["product_id"],
            "sku": it.get("sku") or p.get("sku", ""),
            "product_name": it.get("name") or p.get("name", ""),
            "category": p.get("category", ""),
            "requested_qty": int(pricing["recommended_qty"]) or 1,
            "unit_price": pricing["base_price"],
            "discount_percent": pricing["discount_percent"],
            "margin_percent": it.get("margin_percent"),
            "mrp": pricing["mrp"],
            "selling_price": pricing["selling_price"],
            "line_total": round(pricing["selling_price"] * (int(pricing["recommended_qty"]) or 1), 2),
        })
    return {
        "franchise_id": fid,
        "franchise_name": fr.get("name", ""),
        "tier_id": snap.get("tier_id"),
        "tier_name": snap.get("tier_name", ""),
        "snapshot_version": snap.get("version", 1),
        "default_margin_percent": tier_margin,
        "line_items": lines,
    }


# ============================================================
#  2. OCR ALIAS LEARNING ENGINE
# ============================================================
@router.get("/ocr-aliases", tags=["ocr"])
async def list_aliases(vendor_id: Optional[str] = None,
                        user: dict = Depends(require_roles("super_admin", "hub_accountant", "warehouse_manager"))):
    q = {}
    if vendor_id:
        q["vendor_id"] = vendor_id
    docs = await _db.ocr_aliases.find(q, {"_id": 0}).sort("hits", -1).to_list(500)
    return docs


class AliasLearnRequest(BaseModel):
    vendor_id: Optional[str] = None
    vendor_alias: str
    product_id: str


@router.post("/ocr-aliases/learn", tags=["ocr"])
async def learn_alias(body: AliasLearnRequest, request: Request,
                       user: dict = Depends(require_roles("super_admin", "hub_accountant", "warehouse_manager"))):
    if not body.vendor_alias.strip() or not body.product_id:
        raise HTTPException(400, "vendor_alias and product_id required")
    prod = await _db.products.find_one({"id": body.product_id}, {"_id": 0})
    if not prod:
        raise HTTPException(404, "Product not found")
    alias = (body.vendor_alias or "").strip().upper()
    existing = await _db.ocr_aliases.find_one(
        {"vendor_id": body.vendor_id, "vendor_alias": alias, "product_id": body.product_id},
        {"_id": 0},
    )
    if existing:
        await _db.ocr_aliases.update_one(
            {"id": existing["id"]},
            {"$inc": {"hits": 1}, "$set": {"last_used_at": now_iso()}},
        )
        return {"ok": True, "id": existing["id"], "hits": existing["hits"] + 1}
    rec = OcrAlias(
        vendor_id=body.vendor_id, vendor_alias=alias,
        product_id=body.product_id, sku=prod["sku"],
    )
    await _db.ocr_aliases.insert_one(rec.model_dump())
    await _log_audit(user, "ocr.alias_learn", "ocr_alias", rec.id,
                     after={"alias": alias, "sku": prod["sku"]}, request=request)
    return {"ok": True, "id": rec.id, "hits": 1}


@router.delete("/ocr-aliases/{aid}", tags=["ocr"])
async def delete_alias(aid: str, request: Request,
                        user: dict = Depends(require_roles("super_admin", "hub_accountant"))):
    res = await _db.ocr_aliases.delete_one({"id": aid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Not found")
    await _log_audit(user, "ocr.alias_delete", "ocr_alias", aid, request=request)
    return {"ok": True}


async def lookup_alias(vendor_id: Optional[str], vendor_alias: str) -> Optional[dict]:
    """Used by OCR upload flow — returns matched product dict if alias is known."""
    if not vendor_alias:
        return None
    alias = vendor_alias.strip().upper()
    # Try vendor-specific first
    rec = await _db.ocr_aliases.find_one({"vendor_id": vendor_id, "vendor_alias": alias}, {"_id": 0})
    if not rec:
        # Global fallback (any vendor)
        rec = await _db.ocr_aliases.find_one({"vendor_alias": alias}, {"_id": 0})
    if not rec:
        return None
    prod = await _db.products.find_one({"id": rec["product_id"]}, {"_id": 0})
    if prod:
        # Bump hits/last_used asynchronously (best-effort)
        await _db.ocr_aliases.update_one(
            {"id": rec["id"]}, {"$inc": {"hits": 1}, "$set": {"last_used_at": now_iso()}}
        )
    return prod


# ============================================================
#  3. MULTI-SOURCE ORDERING — Photo + Excel
# ============================================================
ALLOWED_PHOTO = {"image/jpeg", "image/png", "image/jpg", "image/webp"}


async def _build_indent_from_lines(franchise_id: str, raw_lines: List[dict], user: dict,
                                    notes: str, priority: str,
                                    source: str, attachment_url: Optional[str],
                                    request: Request) -> Indent:
    franchise = await _db.franchises.find_one({"id": franchise_id}, {"_id": 0})
    if not franchise:
        raise HTTPException(404, "Franchise not found")
    # franchise tier
    tier = None
    if franchise.get("tier_id"):
        tier = await _db.franchise_tiers.find_one({"id": franchise["tier_id"]}, {"_id": 0})

    items: List[IndentLineItem] = []
    unmatched: List[dict] = []
    total = 0.0
    for r in raw_lines:
        sku = (r.get("sku") or "").strip()
        desc = (r.get("description") or "").strip()
        qty = int(r.get("qty") or 0)
        if qty <= 0:
            continue
        prod = None
        if sku:
            prod = await _db.products.find_one({"sku": {"$regex": f"^{sku}$", "$options": "i"}}, {"_id": 0})
        if not prod and desc:
            # try a name match
            prod = await _db.products.find_one(
                {"name": {"$regex": desc[:25], "$options": "i"}}, {"_id": 0},
            )
        if not prod:
            unmatched.append({"sku": sku, "description": desc, "qty": qty})
            continue
        # tier-aware pricing
        if tier:
            margin = _resolve_margin(tier, prod.get("category", ""), prod.get("margin_percent", 22))
            price = round((prod.get("landing_price", 0) or 0) * (1 + margin / 100), 2)
        else:
            price = prod.get("franchise_price", 0) or 0
        items.append(IndentLineItem(
            product_id=prod["id"], product_name=prod["name"], sku=prod["sku"],
            requested_qty=qty, unit_price=price, line_total=round(price * qty, 2),
        ))
        total += price * qty

    if not items:
        raise HTTPException(400, "No matching products found. Adjust the input and retry.")

    num = await _gen_sequence("indent", "IND")
    indent = Indent(
        indent_number=num, franchise_id=franchise_id, franchise_name=franchise["name"],
        priority=priority, line_items=items, total_amount=round(total, 2),
        notes=notes, created_by=user["id"],
        source=source, source_attachment_url=attachment_url,
    )
    await _db.indents.insert_one(indent.model_dump())
    await _log_audit(user, f"indent.create.{source}", "indent", indent.id,
                     after={"franchise": franchise["name"], "items": len(items),
                            "source": source, "unmatched": len(unmatched)},
                     request=request)
    payload = indent.model_dump()
    payload["unmatched"] = unmatched
    return payload


@router.post("/indents/photo", tags=["indents"])
async def create_indent_from_photo(
    request: Request,
    file: UploadFile = File(...),
    franchise_id: str = Form(...),
    priority: str = Form("routine"),
    notes: str = Form(""),
    user: dict = Depends(get_current_user),
):
    if user["role"] == "franchise_manager" and user.get("franchise_id") != franchise_id:
        raise HTTPException(403, "Cannot create order for another franchise")
    mime = (file.content_type or "").lower()
    if mime not in ALLOWED_PHOTO:
        raise HTTPException(400, f"Unsupported image type: {mime}")
    save_path, public_url = await _save_uploaded_file(file, prefix="order")
    ocr_result = await _ocr_service.parse_photo_order(save_path, mime)
    raw_lines = ocr_result.get("items", [])
    if not raw_lines:
        raise HTTPException(422, ocr_result.get("_error") or "OCR did not find any order lines")
    payload = await _build_indent_from_lines(
        franchise_id=franchise_id, raw_lines=raw_lines, user=user,
        notes=notes or "Created from photo order", priority=priority,
        source="photo", attachment_url=public_url, request=request,
    )
    return {"indent": payload, "ocr_provider": ocr_result.get("provider"), "ocr_model": ocr_result.get("model")}


@router.post("/indents/excel", tags=["indents"])
async def create_indent_from_excel(
    request: Request,
    file: UploadFile = File(...),
    franchise_id: str = Form(...),
    priority: str = Form("routine"),
    notes: str = Form(""),
    user: dict = Depends(get_current_user),
):
    if user["role"] == "franchise_manager" and user.get("franchise_id") != franchise_id:
        raise HTTPException(403, "Cannot create order for another franchise")
    name = (file.filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls") or name.endswith(".csv")):
        raise HTTPException(400, "Upload an .xlsx, .xls or .csv file")
    save_path, public_url = await _save_uploaded_file(file, prefix="order")
    raw_lines = _parse_order_excel(save_path)
    if not raw_lines:
        raise HTTPException(422, "Could not read any SKU/Qty rows from the file")
    payload = await _build_indent_from_lines(
        franchise_id=franchise_id, raw_lines=raw_lines, user=user,
        notes=notes or "Created from Excel order", priority=priority,
        source="excel", attachment_url=public_url, request=request,
    )
    return {"indent": payload}


def _parse_order_excel(path: str) -> List[dict]:
    """Parse uploaded order file → [{sku, qty}]. Tolerant column matching."""
    import pandas as pd
    try:
        if path.lower().endswith(".csv"):
            df = pd.read_csv(path)
        else:
            df = pd.read_excel(path)
    except Exception as e:
        raise HTTPException(422, f"Could not read file: {e}")
    if df.empty:
        return []
    df.columns = [str(c).strip().lower() for c in df.columns]
    sku_col = next((c for c in df.columns if c in {"sku", "code", "item_code", "part_no", "part number"}), None)
    qty_col = next((c for c in df.columns if c in {"qty", "quantity", "nos", "count"}), None)
    if not sku_col or not qty_col:
        # fall back: first non-empty col = sku, second = qty
        non_empty = [c for c in df.columns if df[c].notna().any()]
        if len(non_empty) < 2:
            return []
        sku_col, qty_col = non_empty[0], non_empty[1]
    rows: List[dict] = []
    for _, row in df.iterrows():
        try:
            sku = str(row[sku_col]).strip()
            qty = int(float(row[qty_col]))
        except Exception:
            continue
        if not sku or sku.lower() == "nan" or qty <= 0:
            continue
        rows.append({"sku": sku, "qty": qty})
    return rows


# ============================================================
#  4. BULK INVENTORY IMPORT
# ============================================================
TEMPLATE_COLUMNS = [
    "SKU", "Part Number", "OEM Number", "Product Name", "Category", "HSN",
    "Barcode", "Rack Location", "Vendor", "Landing Price", "MRP",
    "Opening Stock", "Reorder Qty", "Safety Stock",
]


@router.get("/inventory/template", tags=["inventory"])
async def download_inventory_template(user: dict = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for idx, col in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(15, len(col) + 4)
    # Example row
    ws.append([
        "SPK-NEW-001", "PLS150-SPK-01", "OEM-PLS-001",
        "Bajaj Pulsar 150 Spark Plug (Example)",
        "Engine Parts", "8511", "8901234567890",
        "Rack A - Shelf 2 - Bin 5", "Bajaj Auto Genuine Parts Pvt Ltd",
        250, 380, 25, 60, 20,
    ])
    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "Servall ERP — Bulk Inventory Import Template",
        "",
        "Required columns:  SKU, Product Name, Category, HSN, Landing Price, MRP, Opening Stock",
        "Optional columns:  Part Number, OEM Number, Barcode, Rack Location, Vendor, Reorder Qty, Safety Stock",
        "",
        "Notes:",
        "1. SKU must be unique. Duplicates will be flagged in preview.",
        "2. Vendor must match the exact vendor name shown in /vendors page (case-insensitive).",
        "3. HSN must be 4–8 digits.",
        "4. Landing Price = cost price. Franchise & retail prices will auto-compute from tier margin.",
        "5. Opening Stock = quantity at hub-main on import.",
    ]
    for line in instructions:
        ws2.append([line])
    ws2.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_import_template.xlsx"},
    )


async def _validate_inventory_rows(rows: List[dict]) -> List[dict]:
    """Validate each row → attach errors[]. Async because we hit DB for SKU+vendor checks."""
    vendors = await _db.vendors.find({}, {"_id": 0, "id": 1, "name": 1, "code": 1}).to_list(2000)
    vendor_by_name = {(v.get("name") or "").strip().lower(): v for v in vendors}
    vendor_by_code = {(v.get("code") or "").strip().lower(): v for v in vendors}

    # ---- First pass: derive sku list (handle auto-gen) so we can batch existence check
    derived: list[dict] = []
    auto_sku_counter = 0
    for idx, r in enumerate(rows):
        sku = (str(r.get("SKU") or "")).strip()
        name = (str(r.get("Product Name") or "")).strip()
        if not name and not sku:
            continue
        if not sku and name:
            base = "".join(ch for ch in name.upper() if ch.isalnum() or ch == "-")[:24] or "ITEM"
            auto_sku_counter += 1
            sku = f"{base}-{auto_sku_counter:05d}"
            derived.append({"idx": idx, "row": r, "sku": sku, "auto_sku": True})
        else:
            derived.append({"idx": idx, "row": r, "sku": sku, "auto_sku": False})

    # Batch lookup of existing SKUs (single DB round-trip vs N+1)
    all_skus = [d["sku"] for d in derived if d["sku"]]
    existing_skus: set = set()
    if all_skus:
        # Mongo $in supports up to ~1000 efficiently per batch; chunk to be safe
        for i in range(0, len(all_skus), 1000):
            chunk = all_skus[i:i + 1000]
            cursor = _db.products.find({"sku": {"$in": chunk}}, {"_id": 0, "sku": 1})
            async for doc in cursor:
                existing_skus.add(doc["sku"])

    sku_seen: dict = {}
    out: List[dict] = []
    for d in derived:
        idx = d["idx"]
        r = d["row"]
        sku = d["sku"]
        errors: list[str] = []
        warnings: list[str] = []
        if d["auto_sku"]:
            warnings.append(f"SKU auto-generated → {sku}")

        name = (str(r.get("Product Name") or "")).strip()
        category = (str(r.get("Category") or "")).strip()
        hsn = str(r.get("HSN") or "").strip()
        if hsn.endswith(".0"):
            hsn = hsn[:-2]
        try:
            landing = float(r.get("Landing Price") or 0)
        except Exception:
            landing = -1
        try:
            mrp = float(r.get("MRP") or 0)
        except Exception:
            mrp = -1
        try:
            opening = int(float(r.get("Opening Stock") or 0))
        except Exception:
            opening = -1
        try:
            reorder = int(float(r.get("Reorder Qty") or 0))
        except Exception:
            reorder = 0
        try:
            safety = int(float(r.get("Safety Stock") or 0))
        except Exception:
            safety = 0
        vendor_name = (str(r.get("Vendor") or "")).strip()

        # Truncate runaway product names
        if len(name) > 200:
            warnings.append(f"Product name truncated from {len(name)} → 200 chars")
            name = name[:200].strip()

        if not sku:
            errors.append("SKU is required")
        elif sku in sku_seen:
            errors.append(f"Duplicate SKU in upload (row {sku_seen[sku] + 1})")
        elif sku in existing_skus:
            warnings.append("SKU already exists in DB — will UPDATE")
        if sku:
            sku_seen[sku] = idx
        if not name:
            errors.append("Product Name required")
        if not category:
            category = "Uncategorized"
            warnings.append("Category defaulted to 'Uncategorized'")
        if hsn and (not hsn.isdigit() or not (4 <= len(hsn) <= 8)):
            warnings.append(f"HSN '{hsn}' is not 4–8 digit numeric — will be cleared")
            hsn = ""
        if landing < 0:
            errors.append("Invalid Landing Price")
        if mrp < 0:
            errors.append("Invalid MRP")
        if opening < 0:
            errors.append("Invalid Opening Stock")

        vendor_match = None
        if vendor_name:
            v = vendor_by_name.get(vendor_name.lower()) or vendor_by_code.get(vendor_name.lower())
            if v:
                vendor_match = v
            else:
                warnings.append(f"Vendor '{vendor_name}' not found — will leave unassigned")

        is_update = any(w.startswith("SKU already exists") for w in warnings)
        out.append({
            "row": idx + 1,
            "data": {
                "sku": sku, "name": name, "category": category, "hsn": hsn,
                "part_number": (str(r.get("Part Number") or "")).strip(),
                "oem_number": (str(r.get("OEM Number") or "")).strip(),
                "barcode": (str(r.get("Barcode") or "")).strip(),
                "rack_location": (str(r.get("Rack Location") or "")).strip(),
                "landing_price": max(0, landing),
                "mrp": max(0, mrp),
                "opening_stock": max(0, opening),
                "reorder_qty": max(0, reorder),
                "safety_stock": max(0, safety),
                "vendor_id": vendor_match["id"] if vendor_match else None,
                "vendor_name": vendor_match["name"] if vendor_match else "",
            },
            "errors": errors,
            "warnings": warnings,
            "is_update": is_update,
            "blocking": len(errors) > 0,
        })
    return out


# Column aliases — map common exports (Vyapar, Zoho, Tally) to our internal schema
_COLUMN_ALIASES = {
    "Product Name": ["item name", "item name*", "product name", "product", "name", "description"],
    "SKU": ["sku", "item code", "code", "product code", "item id"],
    "Part Number": ["part number", "part no", "part#", "part_no"],
    "OEM Number": ["oem number", "oem no", "oem"],
    "Category": ["category", "categories", "item category", "group"],
    "HSN": ["hsn", "hsn code", "hsn/sac"],
    "Barcode": ["barcode", "ean", "upc"],
    "Rack Location": ["rack location", "rack", "location", "item location", "warehouse location"],
    "Vendor": ["vendor", "supplier", "vendor name", "primary vendor"],
    "Landing Price": ["landing price", "purchase price", "cost price", "cost", "buying price"],
    "MRP": ["mrp", "sale price", "retail price", "selling price", "list price"],
    "Opening Stock": ["opening stock", "current stock quantity", "current stock", "stock", "qty", "quantity", "stock qty"],
    "Reorder Qty": ["reorder qty", "reorder quantity", "reorder", "reorder level"],
    "Safety Stock": ["safety stock", "minimum stock quantity", "min stock", "minimum stock", "min qty"],
}


def _normalize_row(r: dict) -> dict:
    """Map any incoming column name to our canonical schema using _COLUMN_ALIASES."""
    # Build lower-case key map of the incoming row
    lower_map = {(k or "").strip().lower(): k for k in r.keys()}
    out: dict = {}
    for canonical, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in lower_map:
                out[canonical] = r[lower_map[alias]]
                break
    return out


def _read_inventory_upload(path: str) -> List[dict]:
    import pandas as pd
    if path.lower().endswith(".csv"):
        df = pd.read_csv(path, dtype=str, keep_default_na=False)
    else:
        df = pd.read_excel(path, sheet_name=0, dtype=str)
    df = df.dropna(how="all")
    raw = df.fillna("").to_dict(orient="records")
    return [_normalize_row(r) for r in raw]


@router.post("/inventory/import/validate", tags=["inventory"])
async def validate_import(file: UploadFile = File(...),
                           user: dict = Depends(require_roles("super_admin", "warehouse_manager", "hub_accountant"))):
    name = (file.filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls") or name.endswith(".csv")):
        raise HTTPException(400, "Upload an .xlsx, .xls or .csv file")
    save_path, _public_url = await _save_uploaded_file(file, prefix="invimport")
    try:
        rows = _read_inventory_upload(save_path)
    except Exception as e:
        raise HTTPException(422, f"Cannot read file: {e}")
    validated = await _validate_inventory_rows(rows)
    return {
        "total_rows": len(validated),
        "blocking_rows": sum(1 for r in validated if r["blocking"]),
        "warnings_rows": sum(1 for r in validated if not r["blocking"] and r.get("warnings")),
        "updates": sum(1 for r in validated if r["is_update"] and not r["blocking"]),
        "new": sum(1 for r in validated if not r["is_update"] and not r["blocking"]),
        "detected_columns": list(rows[0].keys()) if rows else [],
        "rows": validated,
        "upload_path": save_path,
    }


class CommitImportRequest(BaseModel):
    upload_path: str
    overwrite_existing: bool = True


@router.post("/inventory/import/commit", tags=["inventory"])
async def commit_import(body: CommitImportRequest, request: Request,
                         user: dict = Depends(require_roles("super_admin", "warehouse_manager", "hub_accountant"))):
    if not os.path.exists(body.upload_path):
        raise HTTPException(400, "Upload file no longer available — please re-upload")
    try:
        rows = _read_inventory_upload(body.upload_path)
    except Exception as e:
        raise HTTPException(422, f"Cannot read file: {e}")
    validated = await _validate_inventory_rows(rows)

    created = 0
    updated = 0
    skipped = 0
    for v in validated:
        if v["blocking"]:
            skipped += 1
            continue
        d = v["data"]
        margin = 22.0
        landing = d["landing_price"]
        product_payload = {
            "name": d["name"], "category": d["category"], "hsn_code": d["hsn"],
            "part_number_oem": d["oem_number"], "part_number_aftermarket": d["part_number"],
            "barcode": d["barcode"], "rack_location": d["rack_location"],
            "landing_price": landing, "mrp": d["mrp"],
            "franchise_price": round(landing * (1 + margin / 100), 2),
            "retail_price": round(landing * (1 + (margin + 8) / 100), 2),
            "safety_stock": d["safety_stock"], "reorder_qty": d["reorder_qty"],
            "primary_vendor_id": d["vendor_id"],
            "active": True,
        }
        existing = await _db.products.find_one({"sku": d["sku"]}, {"_id": 0})
        if existing:
            if not body.overwrite_existing:
                skipped += 1
                continue
            await _db.products.update_one({"id": existing["id"]}, {"$set": product_payload})
            pid = existing["id"]
            updated += 1
        else:
            from models import Product
            p = Product(sku=d["sku"], unit="pcs", **product_payload)
            await _db.products.insert_one(p.model_dump())
            pid = p.id
            created += 1
        # Opening stock at hub
        if d["opening_stock"] > 0:
            current = await _db.stock.find_one(
                {"product_id": pid, "location_type": "hub", "location_id": "hub-main"}, {"_id": 0},
            )
            if current:
                delta = d["opening_stock"] - int(current.get("quantity", 0))
                if delta != 0:
                    await _adjust_stock(pid, "hub", "hub-main", delta,
                                        reason="bulk_import_opening_stock",
                                        user_id=user["id"], user_email=user["email"],
                                        reference_type="bulk_import", reference_id=body.upload_path)
            else:
                await _adjust_stock(pid, "hub", "hub-main", d["opening_stock"],
                                    reason="bulk_import_opening_stock",
                                    user_id=user["id"], user_email=user["email"],
                                    reference_type="bulk_import", reference_id=body.upload_path)

    await _log_audit(user, "inventory.bulk_import", "products", "bulk",
                     after={"created": created, "updated": updated, "skipped": skipped},
                     request=request)
    return {"ok": True, "created": created, "updated": updated, "skipped": skipped}


# ============================================================
#  5. EDITABLE PURCHASE ORDERS
# ============================================================
class POLineIn(BaseModel):
    product_id: str
    quantity: int
    unit_price: float


class POSaveRequest(BaseModel):
    vendor_id: str
    notes: str = ""
    line_items: List[POLineIn]
    status: Optional[str] = None  # draft | sent
    po_date: Optional[str] = None  # YYYY-MM-DD; editable order date
    expected_delivery: Optional[str] = None  # YYYY-MM-DD


async def _materialize_po_lines(items: List[POLineIn]) -> tuple[List[dict], float]:
    out = []
    total = 0.0
    for it in items:
        if it.quantity <= 0:
            continue
        p = await _db.products.find_one({"id": it.product_id}, {"_id": 0})
        if not p:
            continue
        line_total = round(it.unit_price * it.quantity, 2)
        out.append({
            "product_id": p["id"], "product_name": p["name"], "sku": p["sku"],
            "quantity": int(it.quantity), "unit_price": float(it.unit_price),
            "line_total": line_total,
        })
        total += line_total
    return out, round(total, 2)


@router.post("/purchase-orders", tags=["purchase_orders"])
async def create_po(body: POSaveRequest, request: Request,
                     user: dict = Depends(require_roles("super_admin", "warehouse_manager", "hub_accountant"))):
    if not body.line_items:
        raise HTTPException(400, "PO must have at least one line")
    vendor = await _db.vendors.find_one({"id": body.vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(404, "Vendor not found")
    lines, total = await _materialize_po_lines(body.line_items)
    if not lines:
        raise HTTPException(400, "No valid lines after validation")
    po_num = await _gen_sequence("po", "PO")
    po = PurchaseOrder(
        po_number=po_num, vendor_id=body.vendor_id, vendor_name=vendor["name"],
        line_items=[POLineItem(**li) for li in lines], total_amount=total,
        auto_generated=False, notes=body.notes, created_by=user["id"],
        status=(body.status if body.status in {"draft", "sent"} else "draft"),
    )
    po_doc = po.model_dump()
    if body.po_date:
        po_doc["po_date"] = body.po_date
    if body.expected_delivery:
        po_doc["expected_delivery"] = body.expected_delivery
    await _db.purchase_orders.insert_one(po_doc)
    po_doc.pop("_id", None)
    await _log_audit(user, "po.create_manual", "purchase_order", po.id,
                     after={"vendor": vendor["name"], "total": total, "lines": len(lines)},
                     request=request)
    return po_doc


@router.put("/purchase-orders/{poid}", tags=["purchase_orders"])
async def update_po(poid: str, body: POSaveRequest, request: Request,
                     user: dict = Depends(require_roles("super_admin", "warehouse_manager", "hub_accountant"))):
    po = await _db.purchase_orders.find_one({"id": poid}, {"_id": 0})
    if not po:
        raise HTTPException(404, "PO not found")
    if po.get("status") not in {"draft", None}:
        raise HTTPException(409, f"Cannot edit a {po['status']} PO")
    vendor = await _db.vendors.find_one({"id": body.vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(404, "Vendor not found")
    lines, total = await _materialize_po_lines(body.line_items)
    if not lines:
        raise HTTPException(400, "No valid lines after validation")
    update = {
        "vendor_id": body.vendor_id, "vendor_name": vendor["name"],
        "line_items": lines, "total_amount": total, "notes": body.notes,
    }
    if body.status in {"draft", "sent"}:
        update["status"] = body.status
    if body.po_date:
        update["po_date"] = body.po_date
    if body.expected_delivery:
        update["expected_delivery"] = body.expected_delivery
    await _db.purchase_orders.update_one({"id": poid}, {"$set": update})
    await _log_audit(user, "po.edit", "purchase_order", poid,
                     before={"vendor": po.get("vendor_name"), "total": po.get("total_amount")},
                     after={"vendor": vendor["name"], "total": total},
                     request=request)
    return {"ok": True}


@router.get("/purchase-orders/{poid}/pdf", tags=["purchase_orders"])
async def po_pdf(poid: str, user: dict = Depends(get_current_user)):
    po = await _db.purchase_orders.find_one({"id": poid}, {"_id": 0})
    if not po:
        raise HTTPException(404, "PO not found")
    vendor = await _db.vendors.find_one({"id": po["vendor_id"]}, {"_id": 0}) or {}
    pdf_bytes = _render_po_pdf(po, vendor)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{po["po_number"]}.pdf"'},
    )


def _render_po_pdf(po: dict, vendor: dict) -> bytes:
    """Minimal professional PO PDF via reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=20, leading=22, spaceAfter=8)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#374151"))
    label = ParagraphStyle("Lbl", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#6b7280"))
    elements = []

    elements.append(Paragraph("<b>Servall</b>", h1))
    elements.append(Paragraph("Purchase Order", styles["Heading3"]))
    elements.append(Spacer(1, 10))
    meta = [
        [Paragraph("PO Number", label), Paragraph(f"<b>{po['po_number']}</b>", small),
         Paragraph("Date", label), Paragraph(po.get("created_at", "")[:10], small)],
        [Paragraph("Status", label), Paragraph(po.get("status", "draft").upper(), small),
         Paragraph("Auto", label), Paragraph("Yes" if po.get("auto_generated") else "No", small)],
    ]
    t0 = Table(meta, colWidths=[70, 180, 60, 180])
    t0.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#f3f4f6")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fafafa")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(t0)
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("<b>Vendor</b>", styles["Heading4"]))
    elements.append(Paragraph(vendor.get("name", po.get("vendor_name", "")), small))
    elements.append(Paragraph(vendor.get("address", ""), small))
    if vendor.get("gstin"):
        elements.append(Paragraph(f"GSTIN: {vendor['gstin']}", small))
    if vendor.get("contact_phone"):
        elements.append(Paragraph(f"Phone: {vendor['contact_phone']}", small))
    elements.append(Spacer(1, 12))

    # Line items
    head = ["#", "SKU", "Product", "Qty", "Rate", "Amount"]
    body = [head]
    for idx, li in enumerate(po.get("line_items", []), start=1):
        body.append([
            str(idx), li.get("sku", ""),
            Paragraph(li.get("product_name", ""), small),
            str(li.get("quantity", 0)),
            f"{li.get('unit_price', 0):,.2f}",
            f"{li.get('line_total', 0):,.2f}",
        ])
    body.append(["", "", "", "", "Total", f"{po.get('total_amount', 0):,.2f}"])
    tbl = Table(body, colWidths=[20, 80, 220, 40, 60, 70])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BACKGROUND", (-2, -1), (-1, -1), colors.HexColor("#f3f4f6")),
        ("FONTNAME", (-2, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 16))

    if po.get("notes"):
        elements.append(Paragraph("<b>Notes</b>", styles["Heading5"]))
        elements.append(Paragraph(po["notes"], small))
        elements.append(Spacer(1, 8))

    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Authorized Signatory — Servall", small))

    doc.build(elements)
    return buf.getvalue()


# ============================================================
#  6. DATE-FILTERED LIST ENDPOINTS (additive — original endpoints untouched)
#     Frontend uses ?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD
# ============================================================
def _date_query(date_from: Optional[str], date_to: Optional[str], field: str = "created_at") -> dict:
    """Build a Mongo range query for an ISO-formatted date field (string compare works for ISO)."""
    q: dict = {}
    rng: dict = {}
    if date_from:
        rng["$gte"] = f"{date_from}T00:00:00"
    if date_to:
        rng["$lte"] = f"{date_to}T23:59:59"
    if rng:
        q[field] = rng
    return q


@router.get("/filtered/indents", tags=["filters"])
async def filtered_indents(date_from: Optional[str] = None, date_to: Optional[str] = None,
                            status: Optional[str] = None, source: Optional[str] = None,
                            user: dict = Depends(get_current_user)):
    q = _date_query(date_from, date_to)
    if user["role"] == "franchise_manager":
        q["franchise_id"] = user.get("franchise_id")
    if status:
        q["status"] = status
    if source:
        q["source"] = source
    docs = await _db.indents.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return docs


@router.get("/filtered/invoices", tags=["filters"])
async def filtered_invoices(date_from: Optional[str] = None, date_to: Optional[str] = None,
                             user: dict = Depends(get_current_user)):
    q = _date_query(date_from, date_to)
    docs = await _db.invoices.find(q, {"_id": 0, "raw_ocr_text": 0}).sort("created_at", -1).to_list(1000)
    return docs


@router.get("/filtered/purchase-orders", tags=["filters"])
async def filtered_pos(date_from: Optional[str] = None, date_to: Optional[str] = None,
                        status: Optional[str] = None,
                        user: dict = Depends(get_current_user)):
    q = _date_query(date_from, date_to)
    if status:
        q["status"] = status
    docs = await _db.purchase_orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return docs


@router.get("/filtered/delivery-challans", tags=["filters"])
async def filtered_dcs(date_from: Optional[str] = None, date_to: Optional[str] = None,
                        user: dict = Depends(get_current_user)):
    q = _date_query(date_from, date_to)
    if user["role"] == "franchise_manager":
        q["franchise_id"] = user.get("franchise_id")
    docs = await _db.delivery_challans.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return docs


@router.get("/filtered/stock-movements", tags=["filters"])
async def filtered_stock_movements(date_from: Optional[str] = None, date_to: Optional[str] = None,
                                    product_id: Optional[str] = None,
                                    user: dict = Depends(require_roles("super_admin", "warehouse_manager", "hub_accountant"))):
    q = _date_query(date_from, date_to, field="timestamp")
    if product_id:
        q["product_id"] = product_id
    docs = await _db.stock_movements.find(q, {"_id": 0}).sort("timestamp", -1).to_list(1000)
    return docs


@router.get("/filtered/audit-logs", tags=["filters"])
async def filtered_audit_logs(date_from: Optional[str] = None, date_to: Optional[str] = None,
                               action: Optional[str] = None,
                               user: dict = Depends(require_roles("super_admin", "hub_accountant"))):
    q = _date_query(date_from, date_to, field="timestamp")
    if action:
        q["action"] = {"$regex": action, "$options": "i"}
    docs = await _db.audit_logs.find(q, {"_id": 0}).sort("timestamp", -1).limit(1000).to_list(1000)
    return docs


@router.get("/filtered/dashboard-trend", tags=["filters"])
async def filtered_dashboard_trend(date_from: Optional[str] = None, date_to: Optional[str] = None,
                                    user: dict = Depends(get_current_user)):
    """Custom-range indent count trend for dashboard."""
    q = _date_query(date_from, date_to)
    if user["role"] == "franchise_manager":
        q["franchise_id"] = user.get("franchise_id")
    indents = await _db.indents.find(q, {"_id": 0, "created_at": 1}).to_list(5000)
    by_day: dict = {}
    for ind in indents:
        day = (ind.get("created_at") or "")[:10]
        if not day:
            continue
        by_day[day] = by_day.get(day, 0) + 1
    series = [{"date": d, "count": c} for d, c in sorted(by_day.items())]
    return {"series": series, "total": len(indents)}
