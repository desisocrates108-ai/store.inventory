"""V2.2 router — Phase 2 & 3 additive endpoints.

Includes:
  • Professional Invoice PDF (vendor-side purchase invoices)
  • Professional Delivery Challan PDF (with verification QR)
  • WhatsApp share deeplinks (wa.me, no API key needed)
  • Reporting Suite: Inventory Value, Stock Movement, Purchase, Sales (PDF + Excel)
  • Auto-reopen 'awaiting_stock' indents on PO receive

All endpoints are additive and mounted at /api by server.py. No legacy
endpoints are modified.
"""
from __future__ import annotations

import io
import logging
import urllib.parse
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Request, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from auth_utils import require_roles, get_current_user

logger = logging.getLogger(__name__)
router = APIRouter()

# Branding constant for PDFs (rebrand-safe).
ORG_NAME = "SERVALL"

# ---- Dependency injection (wired by server.py) ----
_db = None
_log_audit = None
_adjust_stock = None
_compute_tier_price = None


def init(db, log_audit_fn, adjust_stock_fn, compute_tier_price_fn=None):
    global _db, _log_audit, _adjust_stock, _compute_tier_price
    _db = db
    _log_audit = log_audit_fn
    _adjust_stock = adjust_stock_fn
    _compute_tier_price = compute_tier_price_fn


# ============================================================
#  1. PROFESSIONAL INVOICE PDF
# ============================================================
@router.get("/invoices/{iid}/pdf", tags=["invoices"])
async def invoice_pdf(iid: str, user: dict = Depends(get_current_user)):
    inv = await _db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv:
        raise HTTPException(404, "Invoice not found")
    vendor = await _db.vendors.find_one({"id": inv.get("vendor_id")}, {"_id": 0}) or {}
    pdf = _render_invoice_pdf(inv, vendor)
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="invoice-{inv["invoice_number"]}.pdf"'},
    )


class InvoiceEditIn(BaseModel):
    invoice_date: Optional[str] = None
    invoice_number: Optional[str] = None


@router.patch("/invoices/{iid}", tags=["invoices"])
async def edit_invoice_meta(iid: str, body: InvoiceEditIn, request: Request,
                             user: dict = Depends(require_roles("super_admin", "warehouse_manager", "hub_accountant"))):
    inv = await _db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv:
        raise HTTPException(404, "Invoice not found")
    if inv.get("status") == "committed":
        raise HTTPException(400, "Cannot edit a committed invoice")
    update = {}
    if body.invoice_date is not None:
        update["invoice_date"] = body.invoice_date
    if body.invoice_number is not None:
        update["invoice_number"] = body.invoice_number
    if not update:
        return {"ok": True, "noop": True}
    await _db.invoices.update_one({"id": iid}, {"$set": update})
    await _log_audit(user, "invoice.edit", "invoice", iid, before=inv, after=update, request=request)
    return {"ok": True}


# ============================================================
#  2. DELIVERY CHALLAN PDF (redesigned)
# ============================================================
@router.get("/delivery-challans/{dcid}/pdf", tags=["delivery_challans"])
async def dc_pdf(dcid: str, user: dict = Depends(get_current_user)):
    dc = await _db.delivery_challans.find_one({"id": dcid}, {"_id": 0})
    if not dc:
        raise HTTPException(404, "Delivery challan not found")
    franchise = await _db.franchises.find_one({"id": dc.get("franchise_id")}, {"_id": 0}) or {}
    pdf = _render_dc_pdf(dc, franchise)
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="DC-{dc["dc_number"]}.pdf"'},
    )


# ============================================================
#  3. WHATSAPP SHARE (wa.me deeplink — no API key required)
# ============================================================
def _public_pdf_url(request: Request, kind: str, doc_id: str) -> str:
    """Build a full public URL the user can paste into WhatsApp."""
    base = str(request.base_url).rstrip("/")
    return f"{base}/api/{kind}/{doc_id}/pdf"


@router.get("/whatsapp/share", tags=["whatsapp"])
async def whatsapp_share(kind: str = Query(..., pattern="^(invoice|dc|po)$"),
                          doc_id: str = Query(...),
                          phone: Optional[str] = Query(None, description="Recipient WhatsApp number (E.164 without +)"),
                          request: Request = None,
                          user: dict = Depends(get_current_user)):
    """Return a wa.me deeplink with prefilled message and a public PDF URL.

    Frontend just window.opens() this URL. WhatsApp Web / app handles the send.
    """
    coll_map = {"invoice": "invoices", "dc": "delivery_challans", "po": "purchase_orders"}
    kind_path_map = {"invoice": "invoices", "dc": "delivery-challans", "po": "purchase-orders"}
    doc = await _db[coll_map[kind]].find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, f"{kind} not found")

    if kind == "invoice":
        num = doc.get("invoice_number", "")
        party = doc.get("vendor_name", "")
        title = "Purchase Invoice"
    elif kind == "dc":
        num = doc.get("dc_number", "")
        party = doc.get("franchise_name", "")
        title = "Delivery Challan"
    else:
        num = doc.get("po_number", "")
        party = doc.get("vendor_name", "")
        title = "Purchase Order"

    pdf_url = _public_pdf_url(request, kind_path_map[kind], doc_id)
    total = doc.get("grand_total") or doc.get("total_amount") or 0
    message = (
        f"*Servall — {title}*\n"
        f"Number: {num}\n"
        f"Party: {party}\n"
        f"Amount: ₹{total:,.2f}\n\n"
        f"View / Download PDF:\n{pdf_url}"
    )
    encoded = urllib.parse.quote(message)
    if phone:
        # Normalise: strip spaces, leading +, dashes
        clean = "".join(c for c in phone if c.isdigit())
        url = f"https://wa.me/{clean}?text={encoded}"
    else:
        url = f"https://wa.me/?text={encoded}"
    return {"url": url, "message": message, "pdf_url": pdf_url}


# ============================================================
#  4. REPORTING SUITE
# ============================================================
@router.get("/reports/inventory-value", tags=["reports"])
async def report_inventory_value(format: str = Query("json", pattern="^(json|excel|pdf)$"),
                                  user: dict = Depends(require_roles("super_admin", "hub_accountant", "warehouse_manager"))):
    """Stock-on-hand × landing price per product, by category."""
    products = await _db.products.find({}, {"_id": 0}).to_list(20000)
    rows = []
    grand_qty = 0
    grand_value = 0.0
    for p in products:
        stock_doc = await _db.stock.find_one({"product_id": p["id"], "location_type": "hub"}, {"_id": 0})
        qty = int(stock_doc["quantity"]) if stock_doc else 0
        landing = float(p.get("landing_price", 0) or 0)
        value = qty * landing
        rows.append({
            "sku": p.get("sku", ""), "name": p.get("name", ""),
            "category": p.get("category", "Uncategorized"),
            "qty": qty, "landing_price": landing, "value": round(value, 2),
        })
        grand_qty += qty
        grand_value += value
    rows.sort(key=lambda r: r["value"], reverse=True)
    summary = {"total_skus": len(rows), "total_qty": grand_qty, "total_value": round(grand_value, 2)}
    return _format_report("Inventory Value Report", ["SKU", "Name", "Category", "Qty", "Landing", "Value"],
                          [[r["sku"], r["name"], r["category"], r["qty"], r["landing_price"], r["value"]] for r in rows],
                          summary, format, "inventory_value")


@router.get("/reports/stock-movement", tags=["reports"])
async def report_stock_movement(date_from: Optional[str] = None, date_to: Optional[str] = None,
                                 format: str = Query("json", pattern="^(json|excel|pdf)$"),
                                 user: dict = Depends(require_roles("super_admin", "hub_accountant", "warehouse_manager"))):
    q = _date_range_query(date_from, date_to, "created_at")
    docs = await _db.stock_movements.find(q, {"_id": 0}).sort("created_at", -1).to_list(20000)
    rows = []
    total_in = 0
    total_out = 0
    for d in docs:
        delta = int(d.get("delta", 0))
        if delta > 0:
            total_in += delta
        else:
            total_out += -delta
        rows.append({
            "date": (d.get("created_at", "") or "")[:19].replace("T", " "),
            "sku": d.get("sku", ""), "product": d.get("product_name", ""),
            "delta": delta, "reason": d.get("reason", ""),
            "location": f"{d.get('location_type', '')}/{d.get('location_id', '')}",
            "ref": d.get("ref_id", ""),
        })
    summary = {"total_in": total_in, "total_out": total_out, "net": total_in - total_out, "events": len(rows)}
    return _format_report("Stock Movement Report",
                          ["Date", "SKU", "Product", "Delta", "Reason", "Location", "Ref"],
                          [[r["date"], r["sku"], r["product"], r["delta"], r["reason"], r["location"], r["ref"]] for r in rows],
                          summary, format, "stock_movement")


@router.get("/reports/purchase", tags=["reports"])
async def report_purchase(date_from: Optional[str] = None, date_to: Optional[str] = None,
                           format: str = Query("json", pattern="^(json|excel|pdf)$"),
                           user: dict = Depends(require_roles("super_admin", "hub_accountant", "warehouse_manager"))):
    q = _date_range_query(date_from, date_to, "created_at")
    docs = await _db.invoices.find(q, {"_id": 0}).sort("created_at", -1).to_list(20000)
    rows = []
    grand = 0.0
    for d in docs:
        amt = float(d.get("total_amount", 0) or 0)
        grand += amt
        rows.append({
            "date": (d.get("invoice_date") or d.get("created_at", ""))[:10],
            "invoice": d.get("invoice_number", ""), "vendor": d.get("vendor_name", ""),
            "lines": len(d.get("line_items", [])),
            "cgst": float(d.get("cgst", 0) or 0), "sgst": float(d.get("sgst", 0) or 0),
            "total": round(amt, 2), "status": d.get("status", ""),
        })
    summary = {"invoice_count": len(rows), "total_purchase": round(grand, 2)}
    return _format_report("Purchase Report",
                          ["Date", "Invoice #", "Vendor", "Lines", "CGST", "SGST", "Total", "Status"],
                          [[r["date"], r["invoice"], r["vendor"], r["lines"], r["cgst"], r["sgst"], r["total"], r["status"]] for r in rows],
                          summary, format, "purchase")


@router.get("/reports/sales", tags=["reports"])
async def report_sales(date_from: Optional[str] = None, date_to: Optional[str] = None,
                        format: str = Query("json", pattern="^(json|excel|pdf)$"),
                        user: dict = Depends(require_roles("super_admin", "hub_accountant", "warehouse_manager"))):
    """Sales = dispatched/invoiced delivery challans to franchises."""
    q = _date_range_query(date_from, date_to, "created_at")
    docs = await _db.delivery_challans.find(q, {"_id": 0}).sort("created_at", -1).to_list(20000)
    rows = []
    grand = 0.0
    for d in docs:
        amt = float(d.get("grand_total") or d.get("total_amount") or 0)
        grand += amt
        rows.append({
            "date": (d.get("created_at", "") or "")[:10],
            "dc": d.get("dc_number", ""),
            "franchise": d.get("franchise_name", ""),
            "lines": len(d.get("line_items", [])),
            "cgst": float(d.get("cgst", 0) or 0), "sgst": float(d.get("sgst", 0) or 0),
            "total": round(amt, 2), "status": d.get("status", ""),
        })
    summary = {"dc_count": len(rows), "total_sales": round(grand, 2)}
    return _format_report("Sales Report",
                          ["Date", "DC #", "Franchise", "Lines", "CGST", "SGST", "Total", "Status"],
                          [[r["date"], r["dc"], r["franchise"], r["lines"], r["cgst"], r["sgst"], r["total"], r["status"]] for r in rows],
                          summary, format, "sales")


def _date_range_query(date_from: Optional[str], date_to: Optional[str], field: str = "created_at") -> dict:
    q: dict = {}
    rng: dict = {}
    if date_from:
        rng["$gte"] = f"{date_from}T00:00:00"
    if date_to:
        rng["$lte"] = f"{date_to}T23:59:59"
    if rng:
        q[field] = rng
    return q


def _format_report(title: str, columns: list, data_rows: list, summary: dict,
                    fmt: str, filename: str):
    if fmt == "json":
        return {"title": title, "columns": columns, "rows": data_rows, "summary": summary}
    if fmt == "excel":
        return _render_xlsx_response(title, columns, data_rows, summary, filename)
    return _render_pdf_response(title, columns, data_rows, summary, filename)


def _render_xlsx_response(title: str, columns: list, rows: list, summary: dict, filename: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]

    # Header
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

    # Columns
    for ci, c in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=ci, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.alignment = Alignment(horizontal="center")

    # Data
    for ri, row in enumerate(rows, start=4):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)

    # Summary block at bottom
    summary_row = len(rows) + 5
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    for i, (k, v) in enumerate(summary.items()):
        ws.cell(row=summary_row + 1 + i, column=1, value=k.replace("_", " ").title())
        ws.cell(row=summary_row + 1 + i, column=2, value=v)

    # Auto width (basic)
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "AA"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'})


def _render_pdf_response(title: str, columns: list, rows: list, summary: dict, filename: str):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=18, leading=20, spaceAfter=6)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8)
    elements = [
        Paragraph("<b>Servall</b>", h1),
        Paragraph(title, styles["Heading3"]),
        Spacer(1, 8),
    ]

    # Build table — cap rows in PDF to keep it sane (Excel has the full data)
    PDF_ROW_CAP = 500
    body = [columns] + [[str(v)[:60] for v in r] for r in rows[:PDF_ROW_CAP]]
    if len(rows) > PDF_ROW_CAP:
        body.append([f"... + {len(rows) - PDF_ROW_CAP} more rows (use Excel export)"] + [""] * (len(columns) - 1))

    col_widths = [doc.width / len(columns)] * len(columns)
    tbl = Table(body, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 12))

    # Summary
    elements.append(Paragraph("<b>Summary</b>", styles["Heading4"]))
    for k, v in summary.items():
        elements.append(Paragraph(f"{k.replace('_', ' ').title()}: <b>{v}</b>", small))

    doc.build(elements)
    return StreamingResponse(io.BytesIO(buf.getvalue()), media_type="application/pdf",
                             headers={"Content-Disposition": f'inline; filename="{filename}.pdf"'})


# ============================================================
#  5. AUTO-REOPEN PENDING INDENTS ON PO RECEIVE
# ============================================================
async def reopen_awaiting_stock_indents(po: dict):
    """Called from server.py whenever a PO transitions to 'received'.

    For every line in the PO, find any 'awaiting_stock' or 'partially_fulfilled'
    indent that has a backorder for that SKU and move it back to 'pending' so the
    warehouse manager can re-allocate.
    """
    if not po or not po.get("line_items"):
        return {"reopened": 0}
    skus = list({li.get("sku") for li in po["line_items"] if li.get("sku")})
    if not skus:
        return {"reopened": 0}

    candidates = await _db.indents.find({
        "status": {"$in": ["awaiting_stock", "partially_fulfilled"]},
        "line_items.sku": {"$in": skus},
    }, {"_id": 0}).to_list(500)

    reopened_ids = []
    for ind in candidates:
        has_backorder = any(int(li.get("backorder_qty", 0)) > 0 and li.get("sku") in skus
                            for li in ind.get("line_items", []))
        if not has_backorder:
            continue
        await _db.indents.update_one({"id": ind["id"]}, {"$set": {"status": "pending"}})
        reopened_ids.append(ind["id"])

        # Notify warehouse
        await _db.notifications.insert_one({
            "id": __import__("uuid").uuid4().hex,
            "role": "warehouse_manager",
            "user_id": None,
            "title": "Indent reopened — restock received",
            "body": f"Indent {ind['indent_number']} can now be allocated (PO {po.get('po_number')} received).",
            "ref_type": "indent",
            "ref_id": ind["id"],
            "read": False,
            "created_at": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(),
        })
    logger.info(f"reopen_awaiting_stock_indents: PO {po.get('po_number')} reopened {len(reopened_ids)} indents")
    return {"reopened": len(reopened_ids), "indent_ids": reopened_ids}


@router.get("/indents/reopened-by-po/{poid}", tags=["indents"])
async def reopened_indents_by_po(poid: str, user: dict = Depends(get_current_user)):
    """Diagnostic — show what reopen would do, without changing anything."""
    po = await _db.purchase_orders.find_one({"id": poid}, {"_id": 0})
    if not po:
        raise HTTPException(404, "PO not found")
    skus = list({li.get("sku") for li in po.get("line_items", []) if li.get("sku")})
    candidates = await _db.indents.find({
        "status": {"$in": ["awaiting_stock", "partially_fulfilled"]},
        "line_items.sku": {"$in": skus},
    }, {"_id": 0, "id": 1, "indent_number": 1, "status": 1, "franchise_name": 1}).to_list(500)
    return {"po_number": po.get("po_number"), "skus": skus, "indents": candidates}


# ============================================================
#  RENDER HELPERS — Invoice PDF & DC PDF
# ============================================================
def _render_invoice_pdf(inv: dict, vendor: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=32, bottomMargin=32)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=22, leading=24, textColor=colors.HexColor("#111827"))
    sub = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#6b7280"))
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#374151"))
    label = ParagraphStyle("Lbl", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#9ca3af"), spaceAfter=2)

    elements = []

    # Header band
    header = Table([[
        Paragraph("<b>SERVALL NEXUS</b><br/><font size=8 color='#6b7280'>Hyper-Automated ERP</font>", h1),
        Paragraph("<para align='right'><b>PURCHASE INVOICE</b><br/>"
                  f"<font size=9 color='#6b7280'>{inv.get('invoice_number', '')}</font></para>", sub),
    ]], colWidths=[270, 245])
    header.setStyle(TableStyle([
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, 0), 2, colors.HexColor("#111827")),
    ]))
    elements.append(header)
    elements.append(Spacer(1, 12))

    # Meta strip
    meta = [
        [Paragraph("Invoice Date", label),
         Paragraph(f"<b>{inv.get('invoice_date', '') or inv.get('created_at', '')[:10]}</b>", small),
         Paragraph("Status", label),
         Paragraph(f"<b>{(inv.get('status') or '').upper()}</b>", small)],
        [Paragraph("Confidence", label),
         Paragraph(f"<b>{int((inv.get('confidence_score', 0) or 0) * 100)}%</b>", small),
         Paragraph("OCR Model", label),
         Paragraph(f"{inv.get('ocr_model', 'manual')}", small)],
    ]
    mt = Table(meta, colWidths=[70, 180, 70, 180])
    mt.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#f3f4f6")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fafafa")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(mt)
    elements.append(Spacer(1, 12))

    # Vendor block
    elements.append(Paragraph("<b>Vendor</b>", styles["Heading4"]))
    elements.append(Paragraph(vendor.get("name", inv.get("vendor_name", "")), small))
    if vendor.get("address"):
        elements.append(Paragraph(vendor["address"], small))
    if vendor.get("gstin"):
        elements.append(Paragraph(f"GSTIN: <b>{vendor['gstin']}</b>", small))
    if vendor.get("contact_phone"):
        elements.append(Paragraph(f"Phone: {vendor['contact_phone']}", small))
    elements.append(Spacer(1, 12))

    # Line items
    head = ["#", "SKU", "HSN", "Product", "Qty", "Rate", "GST%", "Amount"]
    body = [head]
    sub_total = 0.0
    for idx, li in enumerate(inv.get("line_items", []), start=1):
        amt = float(li.get("line_total", 0) or 0)
        sub_total += amt
        body.append([
            str(idx), li.get("sku", "") or li.get("item_alias", ""),
            li.get("hsn_code", ""),
            Paragraph(li.get("product_name", ""), small),
            f"{li.get('quantity', 0)}",
            f"{float(li.get('unit_price', 0) or 0):,.2f}",
            f"{float(li.get('gst_percent', 0) or 0):.0f}%",
            f"{amt:,.2f}",
        ])
    cgst = float(inv.get("cgst", 0) or 0)
    sgst = float(inv.get("sgst", 0) or 0)
    igst = float(inv.get("igst", 0) or 0)
    total = float(inv.get("total_amount", 0) or 0)
    body.append(["", "", "", "", "", "", "Sub-Total", f"{sub_total:,.2f}"])
    if cgst:
        body.append(["", "", "", "", "", "", "CGST", f"{cgst:,.2f}"])
    if sgst:
        body.append(["", "", "", "", "", "", "SGST", f"{sgst:,.2f}"])
    if igst:
        body.append(["", "", "", "", "", "", "IGST", f"{igst:,.2f}"])
    body.append(["", "", "", "", "", "", "Grand Total", f"{total:,.2f}"])

    tbl = Table(body, colWidths=[20, 70, 50, 180, 35, 55, 40, 65])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND", (-2, -1), (-1, -1), colors.HexColor("#111827")),
        ("TEXTCOLOR", (-2, -1), (-1, -1), colors.white),
        ("FONTNAME", (-2, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 16))

    # Footer
    elements.append(Paragraph("<font color='#6b7280' size=8>Computer-generated invoice. Verified by Servall OCR engine.</font>", small))
    elements.append(Spacer(1, 24))
    elements.append(Paragraph("Authorized Signatory — Servall", small))

    doc.build(elements)
    return buf.getvalue()


def _render_dc_pdf(dc: dict, franchise: dict) -> bytes:
    import qrcode
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=32, bottomMargin=32)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=22, leading=24, textColor=colors.HexColor("#111827"))
    sub = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#6b7280"))
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#374151"))
    label = ParagraphStyle("Lbl", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#9ca3af"))

    # Generate QR for verification
    qr_payload = dc.get("verification_qr") or f"servall-dc:{dc.get('dc_number', '')}:{dc.get('id', '')}"
    qr = qrcode.QRCode(version=1, box_size=4, border=2)
    qr.add_data(qr_payload)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_buf = io.BytesIO()
    qr_img.save(qr_buf, format="PNG")
    qr_buf.seek(0)

    elements = []
    company_block = f"<b>{ORG_NAME}</b><br/><font size=8 color='#6b7280'>Hub &amp; Spoke Distribution</font>"
    header = Table([[
        Paragraph(company_block, h1),
        Paragraph(f"<para align='right'><b>DELIVERY CHALLAN</b><br/>"
                  f"<font size=9 color='#6b7280'>{dc.get('dc_number', '')}</font><br/>"
                  f"<font size=8 color='#6b7280'>Date: {(dc.get('created_at') or '')[:10]}</font></para>", sub),
    ]], colWidths=[270, 245])
    header.setStyle(TableStyle([
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, 0), 2, colors.HexColor("#7a1f1f")),
    ]))
    elements.append(header)
    elements.append(Spacer(1, 12))

    # 2-col block: Ship To + QR
    ship_to = (
        f"<b>{franchise.get('name', dc.get('franchise_name', ''))}</b><br/>"
        f"{franchise.get('address', '')}<br/>"
        f"{franchise.get('city', '')} {franchise.get('state', '')} {franchise.get('pincode', '')}<br/>"
    )
    if franchise.get("contact_phone"):
        ship_to += f"Phone: {franchise['contact_phone']}<br/>"
    if franchise.get("gstin"):
        ship_to += f"GSTIN: <b>{franchise['gstin']}</b>"

    meta_left = [
        [Paragraph("Ship To", label)],
        [Paragraph(ship_to, small)],
        [Spacer(1, 6)],
        [Paragraph(f"<b>Vehicle:</b> {dc.get('vehicle_number', '—')}", small)],
        [Paragraph(f"<b>Transporter:</b> {dc.get('transporter_name', '—')}", small)],
        [Paragraph(f"<b>LR #:</b> {dc.get('lr_number', '—')} &nbsp;&nbsp; <b>E-way:</b> {dc.get('eway_bill_number', '—')}", small)],
    ]
    left_tbl = Table(meta_left, colWidths=[300])

    qr_image = Image(qr_buf, width=110, height=110)
    qr_cell = Table([
        [Paragraph("<para align='center'><b>Scan to Verify</b></para>", small)],
        [qr_image],
        [Paragraph(f"<para align='center'><font size=7 color='#6b7280'>{dc.get('dc_number', '')}</font></para>", small)],
    ], colWidths=[120])
    qr_cell.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    two_col = Table([[left_tbl, qr_cell]], colWidths=[330, 130])
    two_col.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elements.append(two_col)
    elements.append(Spacer(1, 14))

    # Lines
    head = ["#", "SKU", "Product", "Req", "Allocated", "Rate", "Amount"]
    body = [head]
    sub_total = 0.0
    for idx, li in enumerate(dc.get("line_items", []), start=1):
        amt = float(li.get("line_total", 0) or 0)
        sub_total += amt
        body.append([
            str(idx), li.get("sku", ""),
            Paragraph(li.get("product_name", ""), small),
            str(li.get("requested_qty", 0)),
            str(li.get("allocated_qty", 0)),
            f"{float(li.get('unit_price', 0) or 0):,.2f}",
            f"{amt:,.2f}",
        ])
    cgst = float(dc.get("cgst", 0) or 0)
    sgst = float(dc.get("sgst", 0) or 0)
    igst = float(dc.get("igst", 0) or 0)
    grand = float(dc.get("grand_total") or dc.get("total_amount") or 0)

    body.append(["", "", "", "", "", "Sub-Total", f"{sub_total:,.2f}"])
    if cgst:
        body.append(["", "", "", "", "", "CGST", f"{cgst:,.2f}"])
    if sgst:
        body.append(["", "", "", "", "", "SGST", f"{sgst:,.2f}"])
    if igst:
        body.append(["", "", "", "", "", "IGST", f"{igst:,.2f}"])
    body.append(["", "", "", "", "", "Grand Total", f"{grand:,.2f}"])

    tbl = Table(body, colWidths=[20, 75, 215, 40, 55, 50, 65])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND", (-2, -1), (-1, -1), colors.HexColor("#111827")),
        ("TEXTCOLOR", (-2, -1), (-1, -1), colors.white),
        ("FONTNAME", (-2, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 18))

    # Received By / Delivered By blocks (matches RASHI sample layout)
    sig_data = [
        [
            Paragraph("<b>Received By</b>", small),
            Paragraph("<b>Delivered By</b>", small),
        ],
        [
            Paragraph(
                "Name: ____________________________<br/><br/>"
                "Comment: _________________________<br/><br/>"
                f"Date: ____________________________<br/><br/>"
                "Signature: ________________________",
                small,
            ),
            Paragraph(
                f"For <b>{ORG_NAME}</b><br/><br/>"
                "Name: ____________________________<br/><br/>"
                f"Date: ____________________________<br/><br/>"
                "Authorised Signatory: _____________",
                small,
            ),
        ],
    ]
    sig_tbl = Table(sig_data, colWidths=[245, 270])
    sig_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f9fafb")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements.append(sig_tbl)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"<font color='#6b7280' size=8>Generated by {ORG_NAME}. Scan QR to verify authenticity.</font>", small))

    doc.build(elements)
    return buf.getvalue()
